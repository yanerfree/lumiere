import json
import uuid

from fastapi import APIRouter, Body, Depends, Query, UploadFile, File
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from starlette.status import HTTP_201_CREATED

from app.core.exceptions import AppError, NotFoundError
from app.core.audit import write_audit_log
from app.deps.auth import require_project_role
from app.deps.db import get_db
from app.models.project import Branch, Project
from app.models.user import User
from app.schemas.case import BatchCaseRequest, CaseResponse, CopyFromBranchRequest, CreateCaseRequest, UpdateCaseRequest
from app.schemas.common import MessageResponse
from app.services import case_service, folder_service, import_service
from app.services.git_service import get_paths, read_file_content

router = APIRouter(prefix="/api/projects/{project_id}/branches/{branch_id}/cases", tags=["cases"])


@router.post("/import")
async def import_cases(
    project_id: uuid.UUID,
    branch_id: uuid.UUID,
    file: UploadFile = File(...),
    sync_delete: bool = Query(default=False, alias="syncDelete"),
    session: AsyncSession = Depends(get_db),
    _: User = Depends(require_project_role("project_admin", "developer", "tester")),
):
    """导入用例文件（支持 .json 和 .xlsx 格式）。

    默认**只增量**（新增 + 更新）。`syncDelete=true` 才会把
    「之前导入过、这次文件里没有」的用例删掉 —— 那是个破坏性动作，
    必须由人明确选，不能是默认。
    """
    filename = file.filename or ""
    if not (filename.endswith(".json") or filename.endswith(".xlsx")):
        raise AppError(code="INVALID_FILE", message="仅接受 .json 或 .xlsx 文件", status_code=400)

    content = await file.read()
    if len(content) > 50 * 1024 * 1024:
        raise AppError(code="FILE_TOO_LARGE", message="文件大小不能超过 50MB", status_code=400)

    if filename.endswith(".json"):
        try:
            data = json.loads(content)
        except json.JSONDecodeError as e:
            raise AppError(
                code="JSON_PARSE_ERROR",
                message=f"JSON 解析失败：第 {e.lineno} 行",
                status_code=400,
                detail=str(e),
            )
        cases_list = data if isinstance(data, list) else data.get("cases", [])
        if not isinstance(cases_list, list):
            raise AppError(code="INVALID_FORMAT", message="JSON 格式不正确，需要数组或 {cases: [...]}", status_code=400)
    else:
        cases_list = _parse_excel_to_cases(content)

    summary = await import_service.import_cases(
        session, branch_id, cases_list, sync_delete=sync_delete)
    await write_audit_log(session, action="import", target_type="case", changes=summary)
    return {"data": summary}


# 导出/导入两头共用的步骤编解码。抽成一对纯函数，是为了能直接做回环测试 ——
# 光断言"源码里出现了分隔符"没用：写在注释里也算数，埋雷时不会红。
STEP_SEP = " → "


def steps_to_text(steps: list) -> str:
    """`[{action, expected}]` → `1. 动作 → 预期`（没有预期就只写动作）。"""
    out = []
    for i, st in enumerate(steps or []):
        if not isinstance(st, dict):
            out.append(f"{i + 1}. {st}")
            continue
        line = f"{i + 1}. {st.get('action', '')}"
        exp = (st.get("expected") or "").strip()
        out.append(f"{line}{STEP_SEP}{exp}" if exp else line)
    return "\n".join(out)


def text_to_steps(text: str) -> list[dict]:
    """上面那个的逆运算。只导 action 的话，导出→导入一圈每步预期就全丢了，
    而人看 Excel 看不出丢了东西。"""
    import re

    steps = []
    for seq, line in enumerate((text or "").split("\n"), 1):
        line = line.strip()
        if not line:
            continue
        line = re.sub(r"^\d+\.\s*", "", line)
        action, sep, expected = line.partition(STEP_SEP)
        step = {"seq": seq, "action": action.strip()}
        if sep and expected.strip():
            step["expected"] = expected.strip()
        steps.append(step)
    return steps


def _parse_excel_to_cases(content: bytes) -> list[dict]:
    """解析导出的 Excel 文件为用例列表（兼容 export/excel 导出格式）。"""
    import io
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    headers = [str(h).strip() if h else "" for h in rows[0]]
    col = {h: i for i, h in enumerate(headers)}

    cases = []
    for row in rows[1:]:
        def get(name, default=""):
            idx = col.get(name)
            if idx is None or idx >= len(row) or row[idx] is None:
                return default
            return str(row[idx]).strip()

        title = get("标题")
        if not title:
            continue

        module = get("模块")
        submodule = get("子模块")
        case_type = get("测试类型", "api").lower()
        priority = get("优先级", "P2")
        tea_id = get("TEA ID") or get("用例ID") or f"excel-{uuid.uuid4().hex[:8]}"

        status_map = {"已自动化": "automated", "待自动化": "pending", "脚本已移除": "script_removed"}
        auto_status_raw = get("自动化状态", "pending")
        auto_status = status_map.get(auto_status_raw, auto_status_raw)

        script_file = get("脚本文件")
        script_func = get("脚本函数")
        script_ref = {}
        if script_file:
            script_ref["file"] = script_file
        if script_func:
            script_ref["func"] = script_func

        steps = text_to_steps(get("测试步骤"))

        cases.append({
            "tea_id": tea_id,
            "title": title,
            "type": case_type,
            "module": module,
            "submodule": submodule,
            "priority": priority,
            "script_ref": script_ref or None,
            "preconditions": get("前置条件") or None,
            "expected_result": get("预期结果") or None,
            "steps": steps or None,
            "remark": get("备注") or None,
        })

    wb.close()
    return cases


@router.post("", status_code=HTTP_201_CREATED)
async def create_case(
    project_id: uuid.UUID,
    branch_id: uuid.UUID,
    body: CreateCaseRequest,
    session: AsyncSession = Depends(get_db),
    _: User = Depends(require_project_role("project_admin", "developer", "tester")),
):
    """手动创建用例"""
    case = await case_service.create_case(session, branch_id, body)
    await write_audit_log(session, action="create", target_type="case", target_id=case.id, target_name=case.title)
    return {
        "data": CaseResponse.model_validate(case, from_attributes=True).model_dump(by_alias=True)
    }


@router.get("/templates")
async def list_templates(
    project_id: uuid.UUID,
    branch_id: uuid.UUID,
    scenario_type: str = Query(default="api", alias="type"),
    session: AsyncSession = Depends(get_db),
    _: User = Depends(require_project_role("project_admin", "developer", "tester", "guest")),
):
    """查询标记为模板的用例场景"""
    cases = await case_service.list_templates(session, branch_id, scenario_type)
    return {
        "data": [
            CaseResponse.model_validate(c, from_attributes=True).model_dump(by_alias=True)
            for c in cases
        ],
    }


@router.get("")
async def list_cases(
    project_id: uuid.UUID,
    branch_id: uuid.UUID,
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=100, alias="pageSize"),
    case_type: str | None = Query(default=None, alias="type"),
    folder_id: uuid.UUID | None = Query(default=None, alias="folderId"),
    priority: str | None = Query(default=None),
    automation_status: str | None = Query(default=None, alias="automationStatus"),
    is_flaky: bool | None = Query(default=None, alias="isFlaky"),
    keyword: str | None = Query(default=None),
    include_deleted: bool = Query(default=False, alias="includeDeleted"),
    review_status: str | None = Query(default=None, alias="reviewStatus"),
    lifecycle_status: str | None = Query(default=None, alias="lifecycleStatus"),
    manual_status: str | None = Query(default=None, alias="manualStatus"),
    ui_status: str | None = Query(default=None, alias="uiStatus"),
    api_status: str | None = Query(default=None, alias="apiStatus"),
    session: AsyncSession = Depends(get_db),
    _: User = Depends(require_project_role("project_admin", "developer", "tester", "guest")),
):
    """用例列表（分页 + 多条件筛选）"""
    cases, total = await case_service.list_cases(
        session, branch_id, page, page_size,
        case_type=case_type, folder_id=folder_id, priority=priority,
        automation_status=automation_status, is_flaky=is_flaky, keyword=keyword,
        include_deleted=include_deleted, review_status=review_status,
        lifecycle_status=lifecycle_status, manual_status=manual_status,
        ui_status=ui_status, api_status=api_status,
    )
    assets = await case_service.list_case_assets(session, [c.id for c in cases])
    data = []
    for c in cases:
        row = CaseResponse.model_validate(c, from_attributes=True).model_dump(by_alias=True)
        row.update(assets.get(c.id, {}))
        data.append(row)
    return {
        "data": data,
        "pagination": {"page": page, "pageSize": page_size, "total": total},
    }



@router.get("/export/excel")
async def export_cases_excel(
    project_id: uuid.UUID,
    branch_id: uuid.UUID,
    session: AsyncSession = Depends(get_db),
    _: User = Depends(require_project_role("project_admin", "developer", "tester", "guest")),
    keyword: str | None = Query(default=None),
    automation_status: str | None = Query(default=None, alias="automationStatus"),
    folder_id: uuid.UUID | None = Query(default=None, alias="folderId"),
    # 页面上的筛选此前一个都没接：筛了「待审核」点导出，导出来的还是全部。
    # 人拿到文件不会发现，因为文件里看不出它本该是 41 条。
    lifecycle_status: str | None = Query(default=None, alias="lifecycleStatus"),
    review_status: str | None = Query(default=None, alias="reviewStatus"),
    ui_status: str | None = Query(default=None, alias="uiStatus"),
    api_status: str | None = Query(default=None, alias="apiStatus"),
    case_ids: str | None = Query(default=None, alias="caseIds"),
):
    """导出用例为 Excel。

    导的是**手动步骤用例**——接口场景和 UI 脚本的正文不在里面
    （它们分别在 api_test_scenarios / scripts，格式也不是表格能装的）。

    范围：勾了行就只导勾的；没勾就按页面当前的筛选导。
    """
    import io
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    if case_ids:
        picked = [uuid.UUID(x) for x in case_ids.split(",") if x.strip()]
        from sqlalchemy import select as _select

        from app.models.case import Case as _Case
        cases = list((await session.execute(
            _select(_Case).where(_Case.id.in_(picked), _Case.deleted_at.is_(None))
            .order_by(_Case.case_code)
        )).scalars().all())
    else:
        cases, _ = await case_service.list_cases(
            session, branch_id, page=1, page_size=10000,
            keyword=keyword, automation_status=automation_status, folder_id=folder_id,
            lifecycle_status=lifecycle_status, review_status=review_status,
            ui_status=ui_status, api_status=api_status,
        )

    # 加载目录映射: folder_id → (模块名, 子模块名)
    from sqlalchemy import select
    from app.models.case import CaseFolder
    folder_result = await session.execute(select(CaseFolder).where(CaseFolder.branch_id == branch_id))
    all_folders = {f.id: f for f in folder_result.scalars().all()}

    def get_folder_names(fid):
        if not fid or fid not in all_folders:
            return "", ""
        folder = all_folders[fid]
        if folder.parent_id and folder.parent_id in all_folders:
            return all_folders[folder.parent_id].name, folder.name
        return folder.name, ""

    wb = Workbook()
    ws = wb.active
    ws.title = "用例列表"

    headers = [
        "用例ID", "标题", "模块", "子模块", "测试类型", "优先级",
        "自动化状态", "来源", "Flaky",
        "前置条件", "测试步骤", "预期结果",
        "脚本文件", "脚本函数", "TEA ID",
        "备注", "创建时间", "更新时间",
    ]

    header_fill = PatternFill(start_color="E6F0FF", end_color="E6F0FF", fill_type="solid")
    header_font = Font(bold=True, size=11)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    status_map = {"automated": "已自动化", "pending": "待自动化", "script_removed": "脚本已移除", "archived": "已归档"}

    for row_idx, c in enumerate(cases, 2):
        steps_text = steps_to_text(c.steps)

        module_name, sub_module_name = get_folder_names(c.folder_id)

        row = [
            c.case_code or "",
            c.title or "",
            module_name,
            sub_module_name,
            (c.type or "").upper(),
            c.priority or "",
            status_map.get(c.automation_status, c.automation_status or ""),
            "导入" if c.source == "imported" else "手动",
            "是" if c.is_flaky else "否",
            c.preconditions or "",
            steps_text,
            c.expected_result or "",
            c.script_ref_file or "",
            c.script_ref_func or "",
            c.tea_id or "",
            c.remark or "",
            c.created_at.strftime("%Y-%m-%d %H:%M") if c.created_at else "",
            c.updated_at.strftime("%Y-%m-%d %H:%M") if c.updated_at else "",
        ]
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    col_widths = [18, 40, 12, 12, 8, 6, 12, 6, 5, 30, 50, 30, 40, 25, 20, 20, 18, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=cases-export.xlsx"},
    )

@router.post("/{case_id}/release-quarantine")
async def release_quarantine(
    project_id: uuid.UUID,
    branch_id: uuid.UUID,
    case_id: uuid.UUID,
    session: AsyncSession = Depends(get_db),
    _: User = Depends(require_project_role("project_admin", "developer", "tester")),
):
    """人工解除 flaky 自动隔离 —— 脚本修好了、环境稳了，不用干等 14 天。

    判定依据（flaky_evidence）保留备查：解除是"我看过依据、认为不必再隔离"，
    不是"把这段历史抹掉"。
    """
    from app.services import flaky_service

    case = await flaky_service.release(session, case_id)
    if case is None:
        raise NotFoundError(code="CASE_NOT_FOUND", message="用例不存在")
    await session.commit()
    await write_audit_log(session, action="release_quarantine", target_type="case",
                          target_id=case.id, target_name=case.title)
    return {"data": {"caseId": str(case.id), "quarantinedUntil": None}}


@router.post("/{case_id}/quarantine")
async def quarantine_case(
    project_id: uuid.UUID,
    branch_id: uuid.UUID,
    case_id: uuid.UUID,
    session: AsyncSession = Depends(get_db),
    _: User = Depends(require_project_role("project_admin", "developer", "tester")),
):
    """人主动隔离 —— 「我知道它不稳，先别让它挡路」。

    自动检测**不会**做这件事：平台替人决定"这条先不跑了"，等于替人决定不查这个问题。
    所以隔离只能由人发起，而且 14 天到期自动回来。
    """
    from app.services import flaky_service

    case = await flaky_service.quarantine(session, case_id)
    if case is None:
        raise NotFoundError(code="CASE_NOT_FOUND", message="用例不存在")
    await session.commit()
    await write_audit_log(session, action="quarantine", target_type="case",
                          target_id=case.id, target_name=case.title)
    return {"data": {"caseId": str(case.id),
                     "quarantinedUntil": case.quarantined_until.isoformat()}}


@router.post("/{case_id}/confirm-expected")
async def confirm_expected(
    project_id: uuid.UUID,
    branch_id: uuid.UUID,
    case_id: uuid.UUID,
    session: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_project_role("project_admin", "developer", "tester")),
):
    """确认「预期结果」这一列 —— P0 两阶段的第二阶段。

    确认之后才允许给这条 P0 挂接口场景和 UI 脚本。人的介入点刻意选得极窄：
    只看「预期结果」一列，一屏二三十条几分钟看完。让人审全文他会疲劳、会盖章。

    之后改动步骤或预期结果会自动作废这次确认 —— 确认的是当时那一版。
    """
    from datetime import datetime, timezone

    from app.core.exceptions import ValidationError
    from app.models.case import Case

    case = await session.get(Case, case_id)
    if case is None or case.branch_id != branch_id:
        raise NotFoundError(code="CASE_NOT_FOUND", message="用例不存在")
    if not (case.expected_result or "").strip() and not case.steps:
        raise ValidationError(code="NOTHING_TO_CONFIRM",
                              message="这条用例还没有步骤和预期结果，没什么可确认的")

    case.expected_confirmed_at = datetime.now(timezone.utc)
    case.expected_confirmed_by = current_user.id
    await session.commit()
    await write_audit_log(session, action="confirm_expected", target_type="case",
                          target_id=case.id, target_name=case.title)
    return {"data": {
        "caseId": str(case.id),
        "expectedConfirmedAt": case.expected_confirmed_at.isoformat(),
    }}


@router.post("/{case_id}/copy", status_code=HTTP_201_CREATED)
async def copy_case(
    project_id: uuid.UUID,
    branch_id: uuid.UUID,
    case_id: uuid.UUID,
    session: AsyncSession = Depends(get_db),
    _: User = Depends(require_project_role("project_admin", "developer", "tester")),
):
    """复制用例（同分支内）"""
    from sqlalchemy import select
    from app.models.case import Case, CaseFolder
    from app.services.import_service import _next_case_code
    source = await case_service.get_case(session, case_id)
    module = ""
    if source.folder_id:
        folder = (await session.execute(
            select(CaseFolder).where(CaseFolder.id == source.folder_id)
        )).scalar_one_or_none()
        if folder:
            module = folder.path.split("/")[0] if folder.path else folder.name
    case_code = await _next_case_code(session, branch_id, module or "COPY")
    copy = Case(
        branch_id=branch_id,
        case_code=case_code,
        title=f"{source.title}（复制）",
        type=source.type,
        folder_id=source.folder_id,
        priority=source.priority,
        preconditions=source.preconditions,
        steps=source.steps,
        expected_result=source.expected_result,
        source="manual",
        automation_status=source.automation_status,
        script_ref_file=source.script_ref_file,
        script_ref_func=source.script_ref_func,
        remark=source.remark,
    )
    session.add(copy)
    await session.flush()
    await session.refresh(copy)
    await write_audit_log(session, action="copy", target_type="case", target_id=copy.id, target_name=copy.title)
    return {"data": CaseResponse.model_validate(copy, from_attributes=True).model_dump(by_alias=True)}


@router.get("/{case_id}")
async def get_case(
    project_id: uuid.UUID,
    branch_id: uuid.UUID,
    case_id: uuid.UUID,
    session: AsyncSession = Depends(get_db),
    _: User = Depends(require_project_role("project_admin", "developer", "tester", "guest")),
):
    """用例详情"""
    case = await case_service.get_case(session, case_id)
    return {
        "data": CaseResponse.model_validate(case, from_attributes=True).model_dump(by_alias=True)
    }


@router.put("/{case_id}")
async def update_case(
    project_id: uuid.UUID,
    branch_id: uuid.UUID,
    case_id: uuid.UUID,
    body: UpdateCaseRequest,
    session: AsyncSession = Depends(get_db),
    _: User = Depends(require_project_role("project_admin", "developer", "tester")),
):
    """更新用例"""
    case = await case_service.update_case(session, case_id, body)
    await write_audit_log(session, action="update", target_type="case", target_id=case.id, target_name=case.title)
    return {
        "data": CaseResponse.model_validate(case, from_attributes=True).model_dump(by_alias=True)
    }


@router.post("/batch")
async def batch_cases(
    project_id: uuid.UUID,
    branch_id: uuid.UUID,
    body: BatchCaseRequest,
    session: AsyncSession = Depends(get_db),
    _: User = Depends(require_project_role("project_admin", "developer", "tester")),
):
    """批量操作用例（移动/归档/取消归档/修改优先级/标记Flaky/彻底删除）"""
    if body.action == "hard_delete":
        result = await case_service.batch_hard_delete(session, body.case_ids)
        await write_audit_log(session, action="hard_delete", target_type="case", changes={"count": len(body.case_ids)})
        return {"data": result}
    result = await case_service.batch_cases(
        session, branch_id,
        action=body.action,
        case_ids=body.case_ids,
        folder_id=body.folder_id,
        priority=body.priority,
    )
    await write_audit_log(session, action=body.action, target_type="case", changes={"count": len(body.case_ids)})
    return {"data": result}


@router.post("/empty-trash")
async def empty_trash(
    project_id: uuid.UUID,
    branch_id: uuid.UUID,
    session: AsyncSession = Depends(get_db),
    _: User = Depends(require_project_role("project_admin", "developer", "tester")),
):
    """清空回收站——彻底删除该分支全部已软删除的用例。

    单独开路由而不复用 /batch：BatchCaseRequest.case_ids 有 min_length=1 校验，
    清空回收站不需要传 ID，放宽该校验会削弱其它 action 的保护。
    """
    result = await case_service.empty_trash(session, branch_id)
    await write_audit_log(
        session, action="empty_trash", target_type="case",
        changes={"count": result["succeeded"]},
    )
    return {"data": result}


@router.delete("/{case_id}")
async def delete_case(
    project_id: uuid.UUID,
    branch_id: uuid.UUID,
    case_id: uuid.UUID,
    session: AsyncSession = Depends(get_db),
    _: User = Depends(require_project_role("project_admin", "developer", "tester")),
):
    """软删除用例（标记 deleted_at）"""
    case = await case_service.get_case(session, case_id)
    await case_service.delete_case(session, case_id)
    await write_audit_log(session, action="delete", target_type="case", target_id=case_id, target_name=case.title)
    return MessageResponse(message="删除成功").model_dump()


@router.post("/copy-from")
async def copy_from_branch(
    project_id: uuid.UUID,
    branch_id: uuid.UUID,
    body: CopyFromBranchRequest,
    session: AsyncSession = Depends(get_db),
    _: User = Depends(require_project_role("project_admin", "developer", "tester")),
):
    """从其他分支复制用例到当前分支（深拷贝）"""
    result = await case_service.copy_cases_from_branch(
        session, branch_id, body.source_branch_id, body.case_ids
    )
    await write_audit_log(session, action="copy_from", target_type="case", changes={"count": result.get("copied", 0)})
    return {"data": result}


@router.get("/{case_id}/script")
async def get_case_script(
    project_id: uuid.UUID,
    branch_id: uuid.UUID,
    case_id: uuid.UUID,
    session: AsyncSession = Depends(get_db),
    _: User = Depends(require_project_role("project_admin", "developer", "tester", "guest")),
):
    """获取用例关联的脚本源码（从 git bare repo 读取）"""
    case = await case_service.get_case(session, case_id)
    if not case.script_ref_file:
        raise AppError(code="NO_SCRIPT", message="该用例未关联脚本文件", status_code=404)

    project = (await session.execute(
        select(Project).where(Project.id == project_id)
    )).scalar_one_or_none()
    if not project or not project.script_base_path:
        raise AppError(code="NO_GIT_CONFIG", message="项目未配置脚本路径", status_code=400)

    branch = (await session.execute(
        select(Branch).where(Branch.id == branch_id)
    )).scalar_one_or_none()
    if not branch or not branch.last_commit_sha:
        raise AppError(code="NO_SYNC", message="分支尚未同步，请先执行 Git 同步", status_code=400)

    paths = get_paths(project.script_base_path, branch.name)
    content = read_file_content(
        paths["bare_repo"], branch.last_commit_sha, case.script_ref_file
    )
    if content is None:
        raise NotFoundError(
            code="SCRIPT_NOT_FOUND",
            message=f"脚本文件不存在: {case.script_ref_file}（commit: {branch.last_commit_sha[:8]}）",
        )

    return {
        "data": {
            "filePath": case.script_ref_file,
            "funcName": case.script_ref_func,
            "commitSha": branch.last_commit_sha,
            "content": content,
            "language": "python",
        }
    }


# ---- 用例目录 ----

folders_router = APIRouter(
    prefix="/api/projects/{project_id}/branches/{branch_id}/folders", tags=["folders"]
)


@folders_router.get("")
async def list_folders(
    project_id: uuid.UUID,
    branch_id: uuid.UUID,
    session: AsyncSession = Depends(get_db),
    _: User = Depends(require_project_role("project_admin", "developer", "tester", "guest")),
):
    """目录树（含用例计数）"""
    tree = await folder_service.list_folder_tree(session, branch_id)
    return {"data": tree}


@folders_router.get("/empty")
async def list_empty_folders(
    project_id: uuid.UUID,
    branch_id: uuid.UUID,
    session: AsyncSession = Depends(get_db),
    _: User = Depends(require_project_role("project_admin", "developer", "tester", "guest")),
):
    """列出可以清掉的空目录（没有用例、也没有子目录）。

    只列不删 —— 空目录不一定是垃圾，可能是人先把结构搭好还没往里放用例。
    让人看清名单再决定，比替他判断安全。
    """
    return {"data": await folder_service.list_empty_folders(session, branch_id)}


@folders_router.post("/prune-empty")
async def prune_empty_folders(
    project_id: uuid.UUID,
    branch_id: uuid.UUID,
    folder_ids: list[uuid.UUID] = Body(..., embed=True, alias="folderIds"),
    session: AsyncSession = Depends(get_db),
    _: User = Depends(require_project_role("project_admin", "developer")),
):
    """按**明确给出的 id 名单**删空目录。

    不接受"删掉所有空目录"这种指令 —— 名单是人在页面上看过并勾过的，
    服务端再各判一次是否真的空。
    """
    n = await folder_service.prune_empty_folders(session, branch_id, folder_ids)
    await session.commit()
    return {"data": {"pruned": n}}


@folders_router.post("", status_code=HTTP_201_CREATED)
async def create_folder(
    project_id: uuid.UUID,
    branch_id: uuid.UUID,
    name: str = Query(..., min_length=1, max_length=100),
    parent_id: uuid.UUID | None = Query(default=None, alias="parentId"),
    session: AsyncSession = Depends(get_db),
    _: User = Depends(require_project_role("project_admin", "developer", "tester")),
):
    """创建模块/子模块目录"""
    folder = await folder_service.create_folder(session, branch_id, name, parent_id)
    return {"data": folder}


@folders_router.delete("/{folder_id}")
async def delete_folder(
    project_id: uuid.UUID,
    branch_id: uuid.UUID,
    folder_id: uuid.UUID,
    session: AsyncSession = Depends(get_db),
    _: User = Depends(require_project_role("project_admin")),
):
    """删除目录（空目录才可删除）"""
    await folder_service.delete_folder(session, folder_id)
    return MessageResponse(message="删除成功").model_dump()


